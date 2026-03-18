"""
Coupang 选品系统 - 数据导出服务
支持: CSV, Excel (XLSX), PDF 格式导出
"""

import io
import csv
import os
import json
from datetime import datetime
from typing import Optional
from utils.logger import get_logger

logger = get_logger()


class DataExporter:
    """数据导出器"""

    @staticmethod
    def export_csv(data: list[dict], columns: list[dict] = None,
                   filename: str = "export") -> io.BytesIO:
        """
        导出为 CSV 格式

        :param data: 数据列表 [{"col1": val1, ...}, ...]
        :param columns: 列定义 [{"key": "col1", "label": "列名1"}, ...]
        :param filename: 文件名（不含扩展名）
        :return: BytesIO 对象
        """
        if not data:
            data = [{}]

        if not columns:
            columns = [{"key": k, "label": k} for k in data[0].keys()]

        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # 写入表头
        writer.writerow([col["label"] for col in columns])

        # 写入数据
        for row in data:
            writer.writerow([
                _format_value(row.get(col["key"], ""))
                for col in columns
            ])

        # 转为 bytes（UTF-8 BOM 以兼容 Excel 中文）
        content = output.getvalue()
        buf = io.BytesIO()
        buf.write(b'\xef\xbb\xbf')  # UTF-8 BOM
        buf.write(content.encode("utf-8"))
        buf.seek(0)
        return buf

    @staticmethod
    def export_excel(data: list[dict], columns: list[dict] = None,
                     sheet_name: str = "Sheet1",
                     filename: str = "export") -> io.BytesIO:
        """
        导出为 Excel (XLSX) 格式

        :param data: 数据列表
        :param columns: 列定义
        :param sheet_name: 工作表名称
        :param filename: 文件名
        :return: BytesIO 对象
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("[Export] openpyxl 未安装，无法导出 Excel")
            return None

        if not data:
            data = [{}]

        if not columns:
            columns = [{"key": k, "label": k} for k in data[0].keys()]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写入数据
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                value = row.get(col["key"], "")
                cell = ws.cell(row=row_idx, column=col_idx, value=_format_value(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 自动调整列宽
        for col_idx, col in enumerate(columns, 1):
            max_len = len(str(col["label"]))
            for row in data[:50]:  # 只检查前50行
                val = str(row.get(col["key"], ""))
                max_len = max(max_len, min(len(val), 50))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

        # 冻结首行
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_pdf(data: list[dict], columns: list[dict] = None,
                   title: str = "数据导出报告",
                   filename: str = "export") -> io.BytesIO:
        """
        导出为 PDF 格式

        :param data: 数据列表
        :param columns: 列定义
        :param title: 报告标题
        :param filename: 文件名
        :return: BytesIO 对象
        """
        try:
            from fpdf import FPDF
        except ImportError:
            logger.error("[Export] fpdf2 未安装，无法导出 PDF")
            return None

        if not data:
            data = [{}]

        if not columns:
            columns = [{"key": k, "label": k} for k in data[0].keys()]

        # 限制 PDF 列数（太多列会溢出）
        max_cols = min(len(columns), 8)
        columns = columns[:max_cols]

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # 尝试加载中文字体
        font_loaded = False
        font_paths = [
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    pdf.add_font("CJK", "", fp, uni=True)
                    pdf.set_font("CJK", size=10)
                    font_loaded = True
                    break
                except Exception:
                    continue

        if not font_loaded:
            pdf.set_font("Helvetica", size=10)

        # 标题
        pdf.set_font_size(16)
        pdf.cell(0, 12, title, ln=True, align="C")
        pdf.set_font_size(9)
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
        pdf.ln(5)

        # 计算列宽
        page_width = pdf.w - 2 * pdf.l_margin
        col_width = page_width / len(columns)

        # 表头
        pdf.set_font_size(9)
        pdf.set_fill_color(79, 70, 229)  # 主题色
        pdf.set_text_color(255, 255, 255)
        for col in columns:
            pdf.cell(col_width, 8, str(col["label"])[:20], border=1, fill=True, align="C")
        pdf.ln()

        # 数据行
        pdf.set_text_color(0, 0, 0)
        pdf.set_font_size(8)
        for i, row in enumerate(data):
            if i % 2 == 0:
                pdf.set_fill_color(245, 245, 250)
            else:
                pdf.set_fill_color(255, 255, 255)

            for col in columns:
                val = str(_format_value(row.get(col["key"], "")))[:30]
                pdf.cell(col_width, 7, val, border=1, fill=True, align="C")
            pdf.ln()

        # 页脚
        pdf.ln(5)
        pdf.set_font_size(8)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 5, f"Total: {len(data)} records | Amazon Visionary Sourcing Tool", align="C")

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf


# ============================================================
# 预定义导出模板
# ============================================================

# 产品列表导出列定义
PRODUCT_COLUMNS = [
    {"key": "asin", "label": "ASIN"},
    {"key": "title", "label": "商品标题"},
    {"key": "brand", "label": "品牌"},
    {"key": "price", "label": "价格 ($)"},
    {"key": "rating", "label": "评分"},
    {"key": "review_count", "label": "评论数"},
    {"key": "bsr_rank", "label": "BSR 排名"},
    {"key": "category", "label": "类目"},
    {"key": "monthly_sales", "label": "月销量"},
    {"key": "revenue", "label": "月收入 ($)"},
    {"key": "fba_fee", "label": "FBA 费用 ($)"},
    {"key": "profit_margin", "label": "利润率 (%)"},
    {"key": "competition_level", "label": "竞争度"},
    {"key": "opportunity_score", "label": "机会评分"},
]

# 分析报告导出列定义
ANALYSIS_COLUMNS = [
    {"key": "dimension", "label": "分析维度"},
    {"key": "score", "label": "评分"},
    {"key": "summary", "label": "摘要"},
    {"key": "recommendation", "label": "建议"},
]

# 利润计算导出列定义
PROFIT_COLUMNS = [
    {"key": "asin", "label": "ASIN"},
    {"key": "title", "label": "商品标题"},
    {"key": "selling_price", "label": "售价 ($)"},
    {"key": "cost_price", "label": "成本 ($)"},
    {"key": "fba_fee", "label": "FBA 费用 ($)"},
    {"key": "referral_fee", "label": "佣金 ($)"},
    {"key": "shipping_cost", "label": "运费 ($)"},
    {"key": "net_profit", "label": "净利润 ($)"},
    {"key": "profit_margin", "label": "利润率 (%)"},
    {"key": "roi", "label": "ROI (%)"},
]


def _format_value(value):
    """格式化导出值"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, float):
        return round(value, 2)
    return value
